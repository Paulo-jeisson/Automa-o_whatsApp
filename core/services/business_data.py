import csv
import re
import unicodedata
from io import TextIOWrapper

from django.core.exceptions import ValidationError
from django.db import transaction

from core.models import BusinessDataRecord, BusinessDataSource


BUSINESS_DATA_STOP_WORDS = {
    'a', 'as', 'ao', 'aos', 'da', 'das', 'de', 'do', 'dos', 'e', 'em',
    'eu', 'meu', 'minha', 'na', 'nas', 'no', 'nos', 'o', 'os', 'para',
    'por', 'qual', 'quais', 'quai', 'que', 'se', 'sou', 'um', 'uma', 'tem',
    'dado', 'dados', 'lista', 'listar', 'mostra', 'mostrar', 'diga',
    'todo', 'toda', 'todos', 'todas',
}


def _searchable(value):
    normalized = unicodedata.normalize('NFKD', str(value or '').casefold())
    return ''.join(char for char in normalized if not unicodedata.combining(char))


def _singular(value):
    return value[:-1] if len(value) > 3 and value.endswith('s') else value


def _query_tokens(value):
    return [
        _singular(_searchable(token))
        for token in re.findall(r'[\wÀ-ÿ]+', str(value or ''))
        if len(token) >= 2
    ][:20]


def search_business_data(*, empresa, query, limit=50):
    """Busca tolerante a linguagem natural, sempre isolada por empresa."""
    tokens = _query_tokens(query)
    if not tokens:
        return [], {}
    sources = list(BusinessDataSource.objects.filter(empresa=empresa, is_active=True))
    requested_columns = {}
    column_tokens = set()
    for source in sources:
        matches = []
        for column in source.ai_visible_columns or []:
            normalized = _singular(_searchable(column))
            if normalized in tokens:
                matches.append(column)
                column_tokens.add(normalized)
        if matches:
            requested_columns[source.pk] = matches
    terms = [
        token for token in tokens
        if token not in BUSINESS_DATA_STOP_WORDS and token not in column_tokens
    ]
    source_ids = requested_columns.keys() if requested_columns else [source.pk for source in sources]
    ranked = []
    records = BusinessDataRecord.objects.filter(
        empresa=empresa, source_id__in=source_ids, source__is_active=True,
    ).select_related('source')
    for record in records.iterator(chunk_size=500):
        haystack = _searchable(
            f'{record.searchable_text} {record.source.name} {record.source.source_filename} '
            f'{" ".join(record.source.ai_visible_columns or [])}'
        )
        score = sum(1 for term in terms if term in haystack)
        if not terms and requested_columns:
            score = 1
        if score:
            ranked.append((score, record.row_number, record))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    if ranked and terms:
        best_score = ranked[0][0]
        ranked = [item for item in ranked if item[0] == best_score]
    return [item[2] for item in ranked[:limit]], requested_columns


def _clean_header(value, position):
    header = str(value or '').strip()
    return header[:120] or f'coluna_{position}'


def _csv_rows(uploaded):
    uploaded.seek(0)
    wrapper = TextIOWrapper(uploaded.file, encoding='utf-8-sig', newline='')
    try:
        sample = wrapper.read(4096)
        wrapper.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
        yield from csv.reader(wrapper, dialect)
    finally:
        wrapper.detach()


def _xlsx_records(uploaded):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValidationError('O suporte a Excel não está instalado no servidor.') from error
    uploaded.seek(0)
    workbook = load_workbook(uploaded, read_only=True, data_only=True)
    try:
        all_headers = []
        records = []
        for sheet in workbook.worksheets:
            rows = iter(sheet.iter_rows(values_only=True))
            raw_headers = next((list(row) for row in rows if any(value not in (None, '') for value in row)), None)
            if raw_headers is None:
                continue
            headers = [_clean_header(value, index + 1) for index, value in enumerate(raw_headers)]
            sheet_records = []
            for values in rows:
                normalized = [('' if value is None else str(value).strip()) for value in values[:len(headers)]]
                normalized += [''] * (len(headers) - len(normalized))
                if any(normalized):
                    sheet_records.append(dict(zip(headers, normalized)))
            if not sheet_records:
                continue
            for header in headers:
                if header.casefold() not in {item.casefold() for item in all_headers}:
                    all_headers.append(header)
            records.extend((sheet.title, data) for data in sheet_records)
            if len(records) > 10000:
                raise ValidationError('A planilha excede o limite de 10.000 registros.')
        if not records:
            raise ValidationError('A planilha não possui linhas de dados além do cabeçalho.')
        headers = ['planilha', *all_headers]
        return headers, [
            (number, {'planilha': sheet_name, **{header: data.get(header, '') for header in all_headers}})
            for number, (sheet_name, data) in enumerate(records, start=2)
        ]
    finally:
        workbook.close()


def _pdf_records(uploaded):
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise ValidationError('O suporte a PDF não está instalado no servidor.') from error
    uploaded.seek(0)
    try:
        reader = PdfReader(uploaded)
        records = []
        for page_number, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or '').strip()
            if text:
                records.append((page_number, {'pagina': str(page_number), 'conteudo': text}))
    except Exception as error:
        raise ValidationError('Não foi possível ler o PDF enviado.') from error
    if not records:
        raise ValidationError('O PDF não possui texto extraível.')
    return ['pagina', 'conteudo'], records


def _text_records(uploaded):
    uploaded.seek(0)
    raw = uploaded.read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError as error:
        raise ValidationError('O documento de texto deve estar codificado em UTF-8.') from error
    chunks = [item.strip() for item in text.splitlines() if item.strip()]
    if not chunks:
        raise ValidationError('O documento não possui texto preenchido.')
    if len(chunks) > 10000:
        raise ValidationError('O documento excede o limite de 10.000 trechos.')
    return ['trecho', 'conteudo'], [
        (number, {'trecho': str(number), 'conteudo': chunk})
        for number, chunk in enumerate(chunks, start=1)
    ]


def read_spreadsheet(uploaded):
    suffix = uploaded.name.rsplit('.', 1)[-1].lower()
    if suffix == 'pdf':
        return _pdf_records(uploaded)
    if suffix in {'txt', 'md', 'json', 'xml', 'html', 'htm'}:
        return _text_records(uploaded)
    if suffix == 'xlsx':
        return _xlsx_records(uploaded)
    iterator = _csv_rows(uploaded)
    rows = iter(iterator)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise ValidationError('A planilha está vazia.') from error
    headers = [_clean_header(value, index + 1) for index, value in enumerate(raw_headers)]
    if len(headers) != len(set(item.casefold() for item in headers)):
        raise ValidationError('A planilha possui nomes de colunas repetidos.')
    records = []
    for row_number, values in enumerate(rows, start=2):
        normalized = [('' if value is None else str(value).strip()) for value in values[:len(headers)]]
        normalized += [''] * (len(headers) - len(normalized))
        if any(normalized):
            records.append((row_number, dict(zip(headers, normalized))))
        if len(records) > 10000:
            raise ValidationError('A planilha excede o limite de 10.000 registros.')
    if not records:
        content = ' | '.join(str(value or '').strip() for value in raw_headers).strip(' |')
        if content:
            return ['conteudo'], [(1, {'conteudo': content})]
        raise ValidationError('O arquivo não possui conteúdo preenchido.')
    return headers, records


@transaction.atomic
def import_business_data(*, empresa, user, name, data_type, uploaded, visible_columns, replace_existing):
    headers, rows = read_spreadsheet(uploaded)
    if not visible_columns:
        visible_columns = headers
    canonical = {item.casefold(): item for item in headers}
    unknown = [item for item in visible_columns if item.casefold() not in canonical]
    if unknown:
        raise ValidationError(f'Colunas não encontradas: {", ".join(unknown)}.')
    allowed = [canonical[item.casefold()] for item in visible_columns]
    existing = BusinessDataSource.objects.filter(empresa=empresa, name__iexact=name).first()
    if existing and not replace_existing:
        raise ValidationError('Já existe uma base com esse nome. Marque a opção de substituição.')
    if existing:
        existing.delete()
    source = BusinessDataSource.objects.create(
        empresa=empresa, name=name.strip(), data_type=data_type,
        source_filename=uploaded.name[:255], columns=headers,
        ai_visible_columns=allowed, row_count=len(rows), imported_by=user,
    )
    BusinessDataRecord.objects.bulk_create([
        BusinessDataRecord(
            empresa=empresa, source=source, row_number=row_number, data=data,
            searchable_text=_searchable(
                ' '.join(str(data.get(column, '')) for column in allowed)
            )[:8000],
        )
        for row_number, data in rows
    ], batch_size=500)
    return source
