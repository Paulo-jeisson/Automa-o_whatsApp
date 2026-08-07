from io import BytesIO
import sys
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from core.models import Atendimento, BusinessDataRecord, BusinessDataSource, EmpresaCliente
from core.services.ai.tools import AIToolExecutor


class BusinessDataTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='empresa-a', password='senha-segura')
        self.other_user = User.objects.create_user(username='empresa-b', password='senha-segura')
        self.company = EmpresaCliente.objects.create(usuario=self.user, nome='Loja A')
        self.other_company = EmpresaCliente.objects.create(usuario=self.other_user, nome='Loja B')
        self.client.force_login(self.user)

    def test_csv_import_exposes_only_authorized_columns(self):
        upload = SimpleUploadedFile(
            'catalogo.csv', 'produto;preço;estoque;custo interno\nTênis Azul;249,90;8;100\n'.encode(),
            content_type='text/csv',
        )
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Catálogo atual', 'data_type': 'PRODUCT', 'spreadsheet': upload,
            'ai_visible_columns': 'produto, preço, estoque', 'replace_existing': 'on',
        })
        self.assertRedirects(response, reverse('dados_negocio'))
        source = BusinessDataSource.objects.get(empresa=self.company)
        record = source.records.get()
        self.assertEqual(source.row_count, 1)
        self.assertEqual(record.visible_data, {'produto': 'Tênis Azul', 'preço': '249,90', 'estoque': '8'})
        self.assertNotIn('100', record.searchable_text)

    def test_xlsx_import_and_ai_search_are_tenant_scoped(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['produto', 'preço'])
        sheet.append(['Café Premium', '35.00'])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile('produtos.xlsx', content.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.client.post(reverse('dados_negocio'), {
            'name': 'Produtos', 'data_type': 'PRODUCT', 'spreadsheet': upload,
            'ai_visible_columns': 'produto, preço', 'replace_existing': 'on',
        })
        other_source = BusinessDataSource.objects.create(
            empresa=self.other_company, name='Privado', data_type='PRODUCT', source_filename='x.csv',
            columns=['produto', 'preço'], ai_visible_columns=['produto', 'preço'], row_count=1,
        )
        BusinessDataRecord.objects.create(
            empresa=self.other_company, source=other_source, row_number=2,
            data={'produto': 'Segredo', 'preço': '999'}, searchable_text='segredo 999',
        )
        attendance = Atendimento.objects.create(empresa=self.company, nome_cliente='Cliente', telefone_cliente='5511999999999')
        executor = AIToolExecutor(atendimento=attendance)
        result = executor.execute('pesquisar_dados_negocio', {'consulta': 'Café Premium'})
        self.assertEqual(result['resultados'][0]['dados']['preço'], '35.00')
        self.assertEqual(executor.execute('pesquisar_dados_negocio', {'consulta': 'Segredo'})['resultados'], [])

    def test_company_cannot_toggle_another_company_source(self):
        source = BusinessDataSource.objects.create(
            empresa=self.other_company, name='Privado', data_type='OTHER', source_filename='x.csv',
        )
        response = self.client.post(reverse('dados_negocio_status', args=[source.pk]))
        self.assertEqual(response.status_code, 404)
        source.refresh_from_db()
        self.assertTrue(source.is_active)

    def test_text_document_is_imported_without_requiring_columns(self):
        upload = SimpleUploadedFile(
            'politica.md',
            'Prazo de entrega: 5 dias úteis.\nTrocas em até 30 dias.'.encode(),
            content_type='text/markdown',
        )
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Políticas', 'data_type': 'OTHER', 'spreadsheet': upload,
            'ai_visible_columns': '', 'replace_existing': 'on',
        })
        self.assertRedirects(response, reverse('dados_negocio'))
        source = BusinessDataSource.objects.get(empresa=self.company, name='Políticas')
        self.assertEqual(source.columns, ['trecho', 'conteudo'])
        self.assertEqual(source.ai_visible_columns, ['trecho', 'conteudo'])
        self.assertEqual(source.row_count, 2)
        self.assertIn('prazo de entrega', source.records.first().searchable_text)

    def test_pdf_is_imported_as_searchable_pages(self):
        class FakePage:
            def __init__(self, text):
                self.text = text

            def extract_text(self):
                return self.text

        fake_module = SimpleNamespace(
            PdfReader=lambda uploaded: SimpleNamespace(pages=[
                FakePage('Catálogo PDF da empresa'),
                FakePage('Garantia de doze meses'),
            ]),
        )
        upload = SimpleUploadedFile(
            'catalogo.pdf', b'%PDF-test', content_type='application/pdf',
        )
        with patch.dict(sys.modules, {'pypdf': fake_module}):
            response = self.client.post(reverse('dados_negocio'), {
                'name': 'Catálogo PDF', 'data_type': 'PRODUCT', 'spreadsheet': upload,
                'ai_visible_columns': '', 'replace_existing': 'on',
            })
        self.assertRedirects(response, reverse('dados_negocio'))
        source = BusinessDataSource.objects.get(empresa=self.company, name='Catálogo PDF')
        self.assertEqual(source.columns, ['pagina', 'conteudo'])
        self.assertEqual(source.ai_visible_columns, ['pagina', 'conteudo'])
        self.assertEqual(source.row_count, 2)
        self.assertTrue(source.records.filter(searchable_text__icontains='garantia').exists())

    def test_unsupported_executable_file_is_rejected(self):
        upload = SimpleUploadedFile('programa.exe', b'MZ', content_type='application/octet-stream')
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Executável', 'data_type': 'OTHER', 'spreadsheet': upload,
            'ai_visible_columns': '', 'replace_existing': 'on',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Formato não suportado')
        self.assertFalse(BusinessDataSource.objects.filter(empresa=self.company).exists())

    def test_single_line_csv_is_saved_as_text_and_todas_means_all_columns(self):
        upload = SimpleUploadedFile(
            'informacao.csv', 'Atendimento somente com agendamento'.encode(),
            content_type='text/csv',
        )
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Dados', 'data_type': 'OTHER', 'spreadsheet': upload,
            'ai_visible_columns': 'todas', 'replace_existing': 'on',
        })
        self.assertRedirects(response, reverse('dados_negocio'))
        source = BusinessDataSource.objects.get(empresa=self.company, name='Dados')
        self.assertEqual(source.columns, ['conteudo'])
        self.assertEqual(source.ai_visible_columns, ['conteudo'])
        self.assertEqual(source.row_count, 1)
        self.assertIn('atendimento somente', source.records.get().searchable_text)

    def test_xlsx_scans_all_sheets_and_skips_header_only_sheet(self):
        workbook = Workbook()
        summary = workbook.active
        summary.title = 'Resumo'
        summary.append(['OS', 'Cliente', 'Total'])
        details = workbook.create_sheet('Financeiro')
        details.append(['OS', 'Cliente', 'Total'])
        details.append(['1001', 'Maria', '850.00'])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            'relatorio_financeiro.xlsx', content.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Relatório financeiro', 'data_type': 'OTHER', 'spreadsheet': upload,
            'ai_visible_columns': 'todas', 'replace_existing': 'on',
        })
        self.assertRedirects(response, reverse('dados_negocio'))
        source = BusinessDataSource.objects.get(empresa=self.company)
        self.assertEqual(source.row_count, 1)
        self.assertEqual(source.records.get().data['planilha'], 'Financeiro')
        self.assertEqual(source.records.get().data['Total'], '850.00')

        attendance = Atendimento.objects.create(
            empresa=self.company, nome_cliente='Cliente', telefone_cliente='5511888000000',
        )
        result = AIToolExecutor(atendimento=attendance).execute(
            'pesquisar_dados_negocio', {'consulta': 'relatório financeiro'},
        )
        self.assertEqual(result['resultados'][0]['base'], 'Relatório financeiro')
        self.assertEqual(result['resultados'][0]['dados']['Total'], '850.00')

    def test_xlsx_with_headers_only_is_not_reported_as_success(self):
        workbook = Workbook()
        workbook.active.append(['OS', 'Cliente', 'Total'])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            'vazio.xlsx', content.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response = self.client.post(reverse('dados_negocio'), {
            'name': 'Vazio', 'data_type': 'OTHER', 'spreadsheet': upload,
            'ai_visible_columns': 'todas', 'replace_existing': 'on',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'não possui linhas de dados além do cabeçalho')
        self.assertFalse(BusinessDataSource.objects.filter(empresa=self.company).exists())

    def test_whatsapp_tool_understands_request_for_column_values(self):
        source = BusinessDataSource.objects.create(
            empresa=self.company, name='dados', data_type='OTHER',
            source_filename='agenda.xlsx', columns=['Nome', 'Telefone'],
            ai_visible_columns=['Nome', 'Telefone'], row_count=2,
        )
        BusinessDataRecord.objects.create(
            empresa=self.company, source=source, row_number=2,
            data={'Nome': 'Paulo', 'Telefone': '1111'}, searchable_text='paulo 1111',
        )
        BusinessDataRecord.objects.create(
            empresa=self.company, source=source, row_number=3,
            data={'Nome': 'Ana', 'Telefone': '2222'}, searchable_text='ana 2222',
        )
        attendance = Atendimento.objects.create(
            empresa=self.company, nome_cliente='Cliente', telefone_cliente='5511888111111',
        )
        result = AIToolExecutor(atendimento=attendance).execute(
            'pesquisar_dados_negocio', {'consulta': 'Nomes nos dados'},
        )
        self.assertEqual(
            [item['dados'] for item in result['resultados']],
            [
                {'Nome': 'Paulo', 'Telefone': '1111'},
                {'Nome': 'Ana', 'Telefone': '2222'},
            ],
        )
        self.assertTrue(all(item['base'] == 'dados' for item in result['resultados']))

    def test_whatsapp_tool_combines_row_identity_and_requested_column(self):
        source = BusinessDataSource.objects.create(
            empresa=self.company, name='Oficina', data_type='OTHER',
            source_filename='agendamentos.xlsx',
            columns=['Nome', 'Marca', 'Ano', 'Problema'],
            ai_visible_columns=['Nome', 'Marca', 'Ano', 'Problema'], row_count=2,
        )
        BusinessDataRecord.objects.create(
            empresa=self.company, source=source, row_number=2,
            data={'Nome': 'Paulo', 'Marca': 'Fiat', 'Ano': '2016', 'Problema': 'Retriuvi'},
            searchable_text='paulo fiat 2016 retriuvi',
        )
        BusinessDataRecord.objects.create(
            empresa=self.company, source=source, row_number=3,
            data={'Nome': 'Ana', 'Marca': 'Ford', 'Ano': '2020', 'Problema': 'Freio'},
            searchable_text='ana ford 2020 freio',
        )
        attendance = Atendimento.objects.create(
            empresa=self.company, nome_cliente='Cliente', telefone_cliente='5511888222222',
        )
        result = AIToolExecutor(atendimento=attendance).execute(
            'pesquisar_dados_negocio',
            {'consulta': 'Qual problema do meu carro? Sou Paulo, dono do Fiat 2016'},
        )
        self.assertEqual(result['resultados'][0]['dados'], {
            'Nome': 'Paulo', 'Marca': 'Fiat', 'Ano': '2016', 'Problema': 'Retriuvi',
        })

    def test_restaurant_menu_is_available_to_whatsapp_ai(self):
        source = BusinessDataSource.objects.create(
            empresa=self.company, name='Cardápio de hoje', data_type='PRODUCT',
            source_filename='cardapio.xlsx', columns=['Prato', 'Preço', 'Custo interno'],
            ai_visible_columns=['Prato', 'Preço'], row_count=2,
        )
        for row, prato, preco in [(2, 'Feijoada', '29.90'), (3, 'Lasanha', '25.00')]:
            BusinessDataRecord.objects.create(
                empresa=self.company, source=source, row_number=row,
                data={'Prato': prato, 'Preço': preco, 'Custo interno': '10.00'},
                searchable_text=f'{prato} {preco}'.casefold(),
            )
        attendance = Atendimento.objects.create(
            empresa=self.company, nome_cliente='Cliente', telefone_cliente='5511888333333',
        )
        result = AIToolExecutor(atendimento=attendance).execute(
            'pesquisar_dados_negocio', {'consulta': 'Quais são os preços do cardápio?'},
        )
        self.assertEqual(
            [item['dados'] for item in result['resultados']],
            [
                {'Prato': 'Feijoada', 'Preço': '29.90'},
                {'Prato': 'Lasanha', 'Preço': '25.00'},
            ],
        )
